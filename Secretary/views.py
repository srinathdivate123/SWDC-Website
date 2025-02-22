from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import cache_control
from django.contrib import messages, auth
from django.contrib.auth.models import User
from authentication.models import Coordinator, Secretary, Volunteer, Activity, GuardianFaculty, currentData, Event
from authentication.commonPasswords import common_passwords
from django.http import HttpResponse, JsonResponse
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
import os
from PyPDF2 import PdfReader, PdfWriter
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape, A3
import PyPDF2
from datetime import datetime, date
from openpyxl import Workbook
import requests, openpyxl, re
from django.conf import settings




@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def CoordReportsView(request):
    return render (request, 'coord_reports.html')



@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def s_contact_us(request):
    return render (request, 's_contact_us.html')



@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def show_events(request):
    secretary = Secretary.objects.get(email=request.user.email)
    cutoff_date = date.today()
    events = []

    if secretary.activity:
        events.extend(Event.objects.filter(activity=secretary.activity, date__gte=cutoff_date))

    if secretary.flagshipEvent:
        events.extend(Event.objects.filter(activity=secretary.flagshipEvent, date__gte=cutoff_date))
    return render (request, 'show_events.html', {'events': events, 'secretary': secretary})




AASHAKIRAN = {
    "CS-L": ["1-16", "17-32", "33-48", "49-64", "65-78"],
    "IT-F": ["1-15", "16-31", "32-47", "48-63", "64-77"]
}




@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def addEventView(request):
    if request.method == 'GET':
        secretary = Secretary.objects.get(email=request.user.email)
        current = currentData.objects.get(index='Current')
        return render (request, 'add-event.html', {'current': current, 'secretary':secretary})
    if request.method == 'POST':
        activity = request.POST.get('activity')
        date_str = request.POST.get('event-date')
        eventDate = datetime.strptime(date_str, '%Y-%m-%d').date()
        startTime = request.POST.get('start-time')
        endTime = request.POST.get('end-time')
        map_link = request.POST.get('map-link')
        description = request.POST.get('description')
        mode = request.POST.get('mode')
        venue = request.POST.get('venue')
        divisions = request.POST.getlist('divisions')

        # Extract coordinates from the map_link using regex
        coordinates = extract_coordinates(map_link)
        if coordinates:
            latitude, longitude = coordinates
        else:
            latitude, longitude = None, None
            messages.error(request, 'Invalid map link')
            return redirect('add-event')

        new_event = Event(
            activity=activity,
            date=eventDate,
            start_time=startTime,
            end_time=endTime,
            map_link=map_link,
            description=description,
            latitude=float(latitude) if latitude else None,
            longitude=float(longitude) if longitude else None,
            isOnline = True if mode == 'online' else False,
            venue = venue,
            divisions = str(divisions)[1:-1].replace("'", "")
        )
        new_event.save()

        #Only for Aashakiran
        if activity == "Aashakiran":
            for div in divisions:
                dept, division, group = div.split('-')
                group = int(group) - 1

                min_roll, max_roll = AASHAKIRAN[dept + '-' + division][group].split('-')
                min_roll = int(min_roll)
                max_roll = int(max_roll)

                volunteers = Volunteer.objects.filter(activity=activity, div=division, dept=dept, roll__range=(min_roll, max_roll))
                for volunteer in volunteers:
                    volunteer.attendance += f"#{eventDate.strftime('%d-%m-%Y')}" + ", "
                    volunteer.save()
            return redirect('sdashboard')

        for div in divisions:
            if div.count('-') == 1:
                dept, division = div.split('-')
                volunteers = Volunteer.objects.filter(activity=activity, div=division, dept=dept)
                for volunteer in volunteers:
                    volunteer.attendance += f"#{eventDate.strftime('%d-%m-%Y')}" + ", "
                    volunteer.save()

            else:
                dept, division, group = div.split('-')
                group = int(group) - 1

                min_roll, max_roll = settings.GROUPS[activity][group].split('-')
                min_roll = int(min_roll)
                max_roll = int(max_roll)

                volunteers = Volunteer.objects.filter(activity=activity, div=division, dept=dept, roll__range=(min_roll, max_roll))
                for volunteer in volunteers:
                    volunteer.attendance += f"#{eventDate.strftime('%d-%m-%Y')}" + ", "
                    volunteer.save()
        messages.success(request, 'A new event was created successfully! You can view it below!')
        return redirect('show_events')






def extract_coordinates(url):
    regex = r'@?([-+]?[\d.]+),([-+]?[\d.]+)'
    expanded_url = expand_url(url)
    matches = re.search(regex, expanded_url)

    if matches:
        latitude = matches.group(1)
        longitude = matches.group(2)
        return latitude, longitude
    else:
        return None




def expand_url(short_url):
    try:
        response = requests.head(short_url, allow_redirects=True)
        return response.url  # Return the expanded URL
    except requests.RequestException as e:
        return short_url




@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def activityDivisions(request):
    if request.method == 'POST':
        secretary = Secretary.objects.get(email=request.user.email)
        divisions = []

        if secretary.activity:
            #Only for Aashakiran
            if secretary.activity == "Aashakiran":
                for div in settings.DIVISIONS[secretary.activity]:
                    cnt = 1
                    for group in AASHAKIRAN[div]:
                        divisions.append(div + '-' + str(cnt))
                        cnt += 1
            #After removing 'if' for Aashakiran, make this 'elif' to 'if' and the code will work fine.
            elif len(settings.GROUPS[secretary.activity]) == 0:
                for div in settings.DIVISIONS[secretary.activity]:
                    divisions.append(div)
            else:
                for div in settings.DIVISIONS[secretary.activity]:
                    cnt = 1
                    for group in settings.GROUPS[secretary.activity]:
                        divisions.append(div + '-' + str(cnt))
                        cnt += 1

        if secretary.flagshipEvent:
            if len(settings.GROUPS[secretary.flagshipEvent]) == 0:
                for div in settings.DIVISIONS[secretary.flagshipEvent]:
                    divisions.append(div)
            else:
                for div in settings.DIVISIONS[secretary.flagshipEvent]:
                    cnt = 1
                    for group in settings.GROUPS[secretary.flagshipEvent]:
                        divisions.append(div + '-' + str(cnt))
                        cnt += 1

        return JsonResponse({'divisions': divisions})
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)





@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def updateEvent(request):
    if request.method == 'GET':
        return redirect('SDashboard')
    if request.method == 'POST':
        idx = int(request.POST.get('id')) - 1
        venue = request.POST.get('venue')
        startTime = request.POST.get('startTime')
        endTime = request.POST.get('endTime')
        mapLink = request.POST.get('mapLink')

        secretary = Secretary.objects.get(email=request.user.email)

        # cutoff_date = date(2025, 1, 4)
        cutoff_date = date.today()
        events = []

        if secretary.activity:
            events.extend(Event.objects.filter(activity=secretary.activity, date__gte=cutoff_date))

        if secretary.flagshipEvent:
            events.extend(Event.objects.filter(activity=secretary.flagshipEvent, date__gte=cutoff_date))

        if venue:
            events[idx].venue = venue
        if startTime:
            events[idx].start_time = startTime
        if endTime:
            events[idx].end_time = endTime
        if mapLink:
            events[idx].map_link = mapLink
        events[idx].save()
    messages.success(request, 'Event was updated succesfully!')
    return redirect('show_events')





@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def deleteEvent(request):
    if request.method == 'GET':
        return redirect('SDashboard')
    if request.method == 'POST':
        idx = int(request.POST.get('id')) - 1
        secretary = Secretary.objects.get(email=request.user.email)

        # cutoff_date = date(2025, 1, 4)
        cutoff_date = date.today()
        events = []

        if secretary.activity:
            events.extend(Event.objects.filter(activity=secretary.activity, date__gte=cutoff_date))

        if secretary.flagshipEvent:
            events.extend(Event.objects.filter(activity=secretary.flagshipEvent, date__gte=cutoff_date))

        if events[idx].divisions:
            divisions = events[idx].divisions.split(',')
        else:
            divisions = []
        eventDate = events[idx].date.strftime('%d-%m-%Y')
        activity = events[idx].activity

        events[idx].delete()

        #Only for Aashakiran
        if activity == "Aashakiran":
            for div in divisions:
                dept, division, group = div.split('-')
                group = int(group) - 1

                min_roll, max_roll = AASHAKIRAN[dept + '-' + division][group].split('-')
                min_roll = int(min_roll)
                max_roll = int(max_roll)

                volunteers = Volunteer.objects.filter(activity=activity, div=division, dept=dept, roll__range=(min_roll, max_roll))
                for volunteer in volunteers:
                    index = volunteer.attendance.find(eventDate)
                    if index != -1:
                        volunteer.attendance = volunteer.attendance[:index-1] + volunteer.attendance[index+12:]
                        volunteer.save()
            return redirect('show_activities')

        for div in divisions:
            if div.count('-') == 1:
                dept, division = div.split('-')
                volunteers = Volunteer.objects.filter(activity=activity, div=division, dept=dept)
                for volunteer in volunteers:
                    index = volunteer.attendance.find(eventDate)
                    if index != -1:
                        volunteer.attendance = volunteer.attendance[:index-1] + volunteer.attendance[index+12:]
                        volunteer.save()
            else:
                dept, division, group = div.split('-')
                group = int(group) - 1

                min_roll, max_roll = settings.GROUPS[activity][group].split('-')
                min_roll = int(min_roll)
                max_roll = int(max_roll)

                volunteers = Volunteer.objects.filter(activity=activity, div=division, dept=dept, roll__range=(min_roll, max_roll))
                for volunteer in volunteers:
                    index = volunteer.attendance.find(eventDate)
                    if index != -1:
                        volunteer.attendance = volunteer.attendance[:index-1] + volunteer.attendance[index+12:]
                        volunteer.save()
        messages.info(request, 'Event was deleted successfully!')
        return redirect('show_events')




@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def download_attendance(request):
    secretary = Secretary.objects.get(email=request.user.email)
    current = currentData.objects.get(index='Current')
    if request.method == 'GET':
        return render (request, 'download_attendance.html', {'secretary': secretary, 'current': current})

    if request.method == 'POST':
        activity_name = request.POST.get('event-name')

        volunteers = Volunteer.objects.filter(activity=activity_name, registered_academic_year=secretary.registered_academic_year, registered_semester=secretary.registered_semester)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Attendance'

        headers = ['Name', 'Email', 'PRN', 'Contact No.', 'Division', 'No. of Sessions Attended', 'Attendance Percentage', 'Report Marks', 'Data Collection Marks'] # If using older codes, remove 5th & 6th fields of headers list

        # Old code for attendance sheet generation

        # all_dates = set()
        # for volunteer in volunteers:
        #     attendance_dates = volunteer.attendance.split(', ')
        #     for date_entry in attendance_dates:
        #         date = date_entry[1:]
        #         all_dates.add(date)

        # sorted_dates = sorted(all_dates, key=lambda x: x)
        # all_dates = []
        # if volunteers:
        #     attendance_dates = volunteers[0].attendance.split(', ')
        #     for date_entry in attendance_dates:
        #         if len(date_entry) > 1:  # Ensure the entry is not empty and valid
        #             date = date_entry[1:]  # Extract the date, assuming it starts with a special character
        #             all_dates.append(date)



        # Updated code for date-wise attendance.......Uncomment the below code, if date-wise attendance is required in future.

        # all_dates = []
        # events = Event.objects.filter(activity=activity_name)

        # for event in events:
        #     evt_date = '-'.join(str(event.date).split('-')[::-1])
        #     if evt_date not in all_dates:
        #         all_dates.append(evt_date)

        # # sorted_dates = sorted(all_dates, key=lambda x: x)
        # headers.extend(all_dates)

        # for col_num, header in enumerate(headers, 1):
        #     sheet.cell(row=1, column=col_num, value=header)

        # for row_num, volunteer in enumerate(volunteers, start=2):
        #     sheet.cell(row=row_num, column=1, value=volunteer.vname)
        #     sheet.cell(row=row_num, column=2, value=volunteer.email)
        #     sheet.cell(row=row_num, column=3, value=volunteer.prn)
        #     sheet.cell(row=row_num, column=4, value=volunteer.contact_num)

        #     attendance_status = {}
        #     attendance = volunteer.attendance
        #     # for entry in attendance_entries:
        #     #     status = 'Present' if entry.startswith('$') else 'Absent'
        #     #     date = entry[1:]
        #     #     attendance_status[date] = status
        #     for entry in all_dates:
        #         idx = attendance.find(entry)

        #         if idx == -1:
        #             status = 'NA'
        #         elif attendance[idx-1] == '$':
        #             status = 'Present'
        #         else:
        #             status = 'Absent'

        #         attendance_status[entry] = status

        #     for col_num, date in enumerate(all_dates, start=5):
        #         sheet.cell(row=row_num, column=col_num, value=attendance_status.get(date, 'Absent'))


        # Latest code to generate count & percentage of volunteer attendance

        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        row_num = 2

        for volunteer in volunteers:
            att = volunteer.attendance.split(', ')[:-1]
            total_sessions = len(att)
            present_sessions = 0

            for evt in att:
                if evt[0] == '$':
                    present_sessions += 1

            presents_value = str(present_sessions) + ' out of ' + str(total_sessions)
            if total_sessions != 0:
                attendance_percent = "{:.2f}%".format((present_sessions / total_sessions) * 100)
            else:
                attendance_percent = "{:.2f}%".format(0)

            sheet.cell(row=row_num, column=1, value=volunteer.vname)
            sheet.cell(row=row_num, column=2, value=volunteer.email)
            sheet.cell(row=row_num, column=3, value=volunteer.prn)
            sheet.cell(row=row_num, column=4, value=volunteer.contact_num)
            sheet.cell(row=row_num, column=5, value=volunteer.dept + '-' + volunteer.div)
            sheet.cell(row=row_num, column=6, value=presents_value)
            sheet.cell(row=row_num, column=7, value=attendance_percent)
            sheet.cell(row=row_num, column=8, value=volunteer.reportFillingMarks)
            sheet.cell(row=row_num, column=9, value=volunteer.dataCollectionMarks)

            row_num += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename={activity_name}_attendance.xlsx'
        workbook.save(response)
        return response





@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def SecDashboardView(request):
    if request.method == "GET":
        secretary = Secretary.objects.get(email=request.user.email)
        current = currentData.objects.get(index='Current')
        coordinators = Coordinator.objects.filter(submitted=1, Secretary=request.user.username, verified=0, registered_academic_year=secretary.registered_academic_year, registered_semester=secretary.registered_semester)
        SS = Activity.objects.filter(registration_enabled=True, flagship_event=False)
        FE = Activity.objects.filter(registration_enabled=True, flagship_event=True)
        return render(request, "sdashboard.html", {"coordinators": coordinators, "secretary": secretary, "SS": SS, "FE": FE, "current": current})
    else:
        secretary = Secretary.objects.get(email=request.user.email)
        current = currentData.objects.get(index='Current')

        if request.POST["activity"]:
            secretary.activity = request.POST["activity"]
            secretary.registered_academic_year = current.AcademicYear
            secretary.registered_semester = current.Semester
            secretary.save()
            return redirect('SDashboard')

        elif request.POST["flagshipEvent"]:
            secretary.flagshipEvent = request.POST["flagshipEvent"]
            secretary.registered_academic_year = current.AcademicYear
            secretary.registered_semester = current.Semester
            secretary.save()
            return redirect('SDashboard')

        email = request.POST["email"]
        coordinator = Coordinator.objects.get(email=email)
        return render(request, "ViewCoord.html", {"coordinator": coordinator})






@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def viewVolunteerAttendanceView(request):
    if request.method == "POST":
        email = request.POST["email"]
        secretary = Secretary.objects.get(email=request.user.email)
        coordinator = Coordinator.objects.get(email=email)
        volunteers = Volunteer.objects.filter(Cordinator=coordinator.cname, activity=secretary.activity, registered_academic_year=secretary.registered_academic_year, registered_semester=secretary.registered_semester)

        data = []
        for volunteer in volunteers:
            att = {}
            raw_attendance = volunteer.attendance
            attendance = ''
            for a in raw_attendance:
                if a == ' ' or a == ',':
                    continue
                attendance += a

            for i in range(0, len(attendance), 11):
                date = attendance[i+1 : i+11]
                if attendance[i] == "$":
                    att[date] = "Present"
                else:
                    att[date] = "Absent"

            sorted_keys = sorted(att.keys(), key=lambda x: datetime.strptime(x, '%d-%m-%Y'))
            sorted_date_dict = {key: att[key] for key in sorted_keys}
            att.clear()
            att.update(sorted_date_dict)
            data.append({volunteer.vname: att})
        return render(request, "view_volunteer_attendance.html", {"data": data, "coordinator": coordinator, "secretary": secretary})
    else:
        return redirect("coord-details")





@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def s_my_activity(request):
    secretary = Secretary.objects.get(email=request.user.email)

    if secretary.activity:
        activity = Activity.objects.get(name = secretary.activity)
    elif secretary.flagshipEvent:
        activity = Activity.objects.get(name = secretary.flagshipEvent)
    else:
        messages.error(request, 'No activity chosen')
        return render(request, 's_my_activity.html', {"error": True})
    volunteers = Volunteer.objects.filter(activity = activity, registered_academic_year = secretary.registered_academic_year,registered_semester = secretary.registered_semester)

    if request.method == "GET":
        not_yet_submitted = 0
        submitted_yet_to_be_verified = 0
        verified = 0
        rejected = 0
        failed_by_coords = 0
        total = 0
        failed_for_not_submitting_report = 0
        for v in volunteers:
            total += 1
            if v.submitted == 0 and v.verified == 0:
                not_yet_submitted += 1
            elif v.submitted == 1 and v.verified == 0:
                submitted_yet_to_be_verified += 1
            elif v.submitted == 1 and v.verified == 1:
                verified += 1
            elif v.submitted == 0 and v.verified == 2:
                rejected += 1
            elif v.submitted == 1 and v.verified == 3:
                failed_by_coords += 1
            elif v.submitted == 0 and v.verified == 3:
                failed_for_not_submitting_report += 1
        v_yet_to_verified = Volunteer.objects.filter(activity = activity, registered_academic_year = secretary.registered_academic_year,registered_semester = secretary.registered_semester, submitted = 1, verified = 0)
        c_yet_to_verify = {}

        for v in v_yet_to_verified:
            if c_yet_to_verify.get(v.Cordinator) is not None:
                c_yet_to_verify[v.Cordinator] += 1
            else:
                c_yet_to_verify[v.Cordinator] = 1
        c_yet_to_verify = dict(sorted(c_yet_to_verify.items()))

        context = {
            'activity' : activity,
            'secretary' : secretary,
            'not_yet_submitted' : not_yet_submitted,
            'submitted_yet_to_be_verified' : submitted_yet_to_be_verified,
            'verified' : verified,
            'rejected' : rejected,
            'failed_by_coords' : failed_by_coords,
            'failed_for_not_submitting_report' : failed_for_not_submitting_report,
            'total' : total,
            'c_yet_to_verify' : c_yet_to_verify,
            'len' : len(c_yet_to_verify),
            'academic_year': secretary.registered_academic_year,
            'sem':secretary.registered_semester
        }
        return render(request,"s_my_activity.html",context)
    if request.method == "POST":
        wb = Workbook()
        ws = wb.active
        ws.title = secretary.activity
        ws.append([
            "Name",
            "Submitted",
            "Verified",
            "Email",
            "Gender",
            "Cordinator",
            "Guardian Faculty",
            "Activity",
            "Department",
            "Registered Academic Year",
            "Semester",
            "Academic Year",
            "Division",
            "Phone Number",
            "Parent's Phone Number",
            "Blood Group",
            "PRN",
            "Roll No.",
            "Current Address",
            "Objective Of The Activity",
            "Description Of The Activity",
            "Benefits To The Society",
            "Benefits To Self",
            "Learning Experiences & Challenges",
            "How Did It Shape Your Empathy",
            "URL",
            "Rejection Reason",
            "Attendance",
        ])

        rows = Volunteer.objects.filter(activity=activity, registered_academic_year=secretary.registered_academic_year, registered_semester=secretary.registered_semester)
        if  len(rows) == 0:
            messages.info(request, "No volunteers are registered for " + activity.name + " currently.")
            return redirect("s_my_activity")
        for v in rows:
            ws.append([v.vname, v.submitted, v.verified, v.email, v.gender, v.Cordinator, v.guardian_faculty, v.activity, v.dept, v.registered_academic_year, v.registered_semester, v.academic_year, v.div, v.contact_num, v.parent_num, v.blood_group, v.prn, v.roll, v.current_add, v.ans1, v.ans2, v.ans3, v.ans4, v.ans5, v.ans6, v.url, v.rejection_reason, v.attendance])

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        fname = secretary.activity + ' - Data.xlsx'
        response['Content-Disposition'] = f'attachment; filename={fname}'
        return response





@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def coordDetailsView(request):
    if request.method == "GET":
        secretary = Secretary.objects.get(email=request.user.email)
        secretaries = []

        if secretary.flagshipEvent:
            flagship_coordinators = Coordinator.objects.filter(flagshipEvent=secretary.flagshipEvent, registered_academic_year=secretary.registered_academic_year, registered_semester=secretary.registered_semester)
            secretaries.extend(Secretary.objects.filter(activity = secretary.flagshipEvent, registered_academic_year=secretary.registered_academic_year, registered_semester=secretary.registered_semester))
        else:
            flagship_coordinators = []

        if secretary.activity:
            social_coordinators = Coordinator.objects.filter(activity=secretary.activity, registered_academic_year=secretary.registered_academic_year, registered_semester=secretary.registered_semester)
            secretaries.extend(Secretary.objects.filter(activity = secretary.activity, registered_academic_year=secretary.registered_academic_year, registered_semester=secretary.registered_semester))
        else:
            social_coordinators = []

        s = []
        countRegistered = 0
        countAttendance = 0

        for coord in flagship_coordinators:
            countRegistered += 1
            if coord.marked_attendance_FE:
                countAttendance += 1
        for coord in social_coordinators:
            countRegistered += 1
            if coord.marked_attendance_GP2:
                countAttendance += 1
        for sec in secretaries:
            if sec.sname != secretary.sname:
                s.append(sec.sname)
        return render(request,"coordDetails.html",{"flagship_coordinators": flagship_coordinators,"social_coordinators": social_coordinators, "secretary": secretary, "secretaries" : s, "countRegistered":countRegistered, "countAttendance":countAttendance})
    else:
        return redirect("SDashboard")






@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def SetpasswordPageView(request):
    if request.method == "GET":
        return render(request, "SPassword.html")
    else:
        password = request.POST["password"]
        password = password.strip()
        if password in common_passwords:
            messages.error(request, "This password is too common. Choose a safer one.")
            return redirect("sresetpass")
        user = User.objects.get(username=request.user.username)
        user.set_password(password)
        user.save()
        secretary = Secretary.objects.get(email=user.email)
        secretary.password_changed = True
        secretary.save()
        auth.logout(request)
        messages.success(
            request, "Your password has been changed successfully! Please login again!"
        )
        return redirect("login")






@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def ApproveCoord(request):
    if request.method == "POST":
        coord = Coordinator.objects.get(email=request.POST["email"])
        coord.verified = 1
        coord.save()
        messages.success(
            request, "Coordinator " + coord.cname + " verified successfully!"
        )
        return redirect("SDashboard")
    else:
        return redirect("SDashboard")






@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def rejectCoordView(request):
    if request.method == "POST":
        coordinator = Coordinator.objects.get(email=request.POST["email"])
        coordinator.verified = 2
        coordinator.submitted = 0
        coordinator.save()
        messages.error(request, "Coordinator " + coordinator.cname + " rejected.")
        return redirect("SDashboard")
    else:
        return redirect("SDashboard")





@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def failVolunteersView(request):
    if request.method == 'GET':
        secretary = Secretary.objects.get(email = request.user.email)

        if secretary.activity:
            activity = Activity.objects.get(name = secretary.activity)
        elif secretary.flagshipEvent:
            activity = Activity.objects.get(name = secretary.flagshipEvent)
        else:
            messages.error(request, 'No activity chosen')
            return render(request, 'failVolunteers.html', {"error": True})

        volunteers = Volunteer.objects.filter(activity = activity, registered_semester = secretary.registered_semester, registered_academic_year = secretary.registered_academic_year, verified = 0)
        return render(request, 'failVolunteers.html', {'volunteers': volunteers, 'secretary':secretary, 'activity': activity})
    else:
        name = request.POST['name']
        reason = request.POST['reason']
        secretary = Secretary.objects.get(email = request.user.email)
        volunteer = Volunteer.objects.get(vname = name)
        volunteer.verified = 3
        volunteer.submitted = 1
        volunteer.rejection_reason = reason
        volunteer.save()
        email_subject = "Social Services Course: Update"
        formatedMsg = []
        formatedMsg.append("The 'Social Services' course required active participation, with successful completion of " + volunteer.activity + " being the primary criterion for passing.")
        formatedMsg.append("Regrettably, we are informing you that you have Failed in the 'Social Services' course.")
        formatedMsg.append("The reason, as stated by " + secretary.sname + ", is:")
        formatedMsg.append('"' + reason + '"')
        context = {"name": volunteer.vname, "messages": formatedMsg}
        email_body = render_to_string("email_template.html", context)
        email = EmailMessage(email_subject, email_body, "noreply@semycolon.com", [volunteer.email])
        email.content_subtype = "html"
        email.send(fail_silently = False)
        messages.info(request,"Volunteer " + volunteer.vname + " has been marked as failed. They have been mailed about it. They cannot submit the report again")
        return redirect('fail-volunteers')






@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def showCertificate(request):
    if request.method == 'GET':
        secretary = Secretary.objects.get(email = request.user.email)
        act = secretary.activity.replace(" ", "_")
        template_path = os.path.join(settings.BASE_DIR, "certificateTemplates") + "/" + act + "/" + act + "_" + secretary.registered_academic_year + "_" + str(secretary.registered_semester) + ".pdf"
        if not os.path.exists(template_path):
            messages.error(request, secretary.activity + '\'s certificate is not available on our servers.')
            return redirect("s_my_activity")
        certificate_template = PdfReader(open(template_path, "rb"))
        packet = io.BytesIO()
        output_buffer = io.BytesIO()
        canvasObj = canvas.Canvas(packet, pagesize=landscape(A3))
        canvasObj.setFont("Times-Bold", 27)
        width, height = A3
        text_width = canvasObj.stringWidth('Volunteer Name', "Times-Roman", 27)
        x_center = (width - text_width) / 2
        canvasObj.drawString(x_center, settings.coordinate[act], 'Volunteer Name')
        canvasObj.save()
        packet.seek(0)
        new_pdf = PdfReader(packet)
        output = PdfWriter()
        page = certificate_template.pages[0]
        page.merge_page(new_pdf.pages[0])
        output.add_page(page)
        output.write(output_buffer)
        output_buffer.seek(0)
        response = HttpResponse(output_buffer, content_type='application/pdf')
        return response
    else:
        messages.error(request, 'Not allowed.')
        return redirect('s_my_activity')





@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def showReport(request):
    if request.method == 'GET':
        template_path = os.path.join(settings.BASE_DIR, "certificateTemplates/ReportTemplate.pdf")
        with open(template_path, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            pdf_writer = PyPDF2.PdfWriter()
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]
                pdf_writer.add_page(page)
            output_pdf = io.BytesIO()
            pdf_writer.write(output_pdf)
            output_pdf.seek(0)
            response = HttpResponse(output_pdf, content_type='application/pdf')
            return response
    else:
        messages.error(request, 'Not allowed.')
        return redirect('s_my_activity')






@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required(login_url="/a/login-restricted")
def ReportFillingSampleView(request):
    if request.method == 'GET':
        guardian_faculties = GuardianFaculty.objects.all()
        user = User.objects.get(email = request.user.email)
        hours = str(user.last_login)[11:13]
        minutes = str(user.last_login)[14:16]
        return render(request, 'report-filling.html', {'guardian_faculties':guardian_faculties, 'hours':hours, 'minutes':minutes})
    else:
        messages.error(request, 'Not allowed.')
        return redirect('s_my_activity')