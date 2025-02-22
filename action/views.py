from django.shortcuts import redirect, render
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from authentication.models import Coordinator, Volunteer, Activity, Domain, currentData, Secretary, stats, Attendance
from itertools import chain
import zipfile
import os
from PyPDF2 import PdfReader, PdfWriter
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape, A3
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from .captcha import FormWithCaptcha
from django.views.decorators.csrf import csrf_exempt
import json
import datetime
import openpyxl
from django.utils.safestring import mark_safe
from datetime import date
from collections import defaultdict
from django.contrib.auth.models import User
from django.conf import settings
from Volunteer.views import upload_to_google_drive

def verifyAdmin(request):
    if request.method == 'POST':
        code = request.POST['code']
        if code == os.getenv('ACTIONS_PAGE_VERIFICATION_CODE'):
            return render(request, 'actions.html')
        else:
            return render(request, 'actions_login_page.html', {'message': 'Invalid code'})
    return render(request, 'actions_login_page.html')


def deleteVolunteer(request):
    if request.method == 'POST':
        email = request.POST['email']
        try:
            user = User.objects.get(email=email)
            volunteer = Volunteer.objects.get(email=email)
            volunteer.delete()
            user.delete()
            return render(request, 'actions.html', {'message': 'Volunteer deleted successfully from both the tables'})
        except:
            return render(request, 'actions.html', {'message': 'Volunteer does not exist'})
    return render(request, 'actions.html')

def onlyVolunteer(request):
    if request.method == 'POST':
        email = request.POST['email']
        try:
            volunteer = Volunteer.objects.get(email=email)
            volunteer.delete()
            return render(request, 'actions.html', {'message': 'Volunteer deleted successfully from Volunteer table'})
        except:
            return render(request, 'actions.html', {'message': 'Volunteer does not exist'})
    return render(request, 'actions.html')

def viewVolunteer(request):
    if request.method == 'POST':
        email = request.POST['email']
        try:
            volunteer = Volunteer.objects.get(email=email)
            return render(request, 'actions.html', {'message': volunteer.vname + ' found'})
        except:
            return render(request, 'actions.html', {'message': 'Volunteer does not exist'})
    return render(request, 'actions.html')

def deleteAttendance(request):
    if request.method == 'POST':
        cutoff_date = date(2024, 12, 31)
        activities_to_delete = Activity.objects.filter(date__lte=cutoff_date)
        activities_to_delete.delete()

        volunteers = Volunteer.objects.all()

        for vol in volunteers:
            if '2025' in vol.attendance:
                vol.attendance = vol.attendance.split('2024, ')[-1]
            else:
                vol.attendance = ''
            vol.save()

        return render(request, 'actions.html', {'message': 'Attendance deleted successfully!'})
    return render(request, 'actions.html')

def coordAttendance(request):
    if request.method == 'POST':
        target_date = datetime(2025, 1, 4)
        attendance_records = Attendance.objects.filter(time__date__gte=target_date.date()).order_by('time')

        datewise_coords = defaultdict(set)
        for record in attendance_records:
            record_date = record.time.date()
            if record.coord_name:  # Only include records with a coordinator name
                datewise_coords[record_date].add(record.coord_name)

        content = "Coordinators Attendance:\n"
        for date, coords in datewise_coords.items():
            content += f"\n{date}:\n"
            content += "\n".join(f"  - {name}" for name in sorted(coords))

        # Generate response with the text file
        response = HttpResponse(content, content_type='text/plain')
        response['Content-Disposition'] = 'attachment; filename="coord_attendance.txt"'
        return response
    return render(request, 'actions.html')


@csrf_exempt
def fetchVolunteerNames(request):
    if request.method == 'GET':
        return JsonResponse({'Identity-Check': 'Oops! You seem to be an intruder!'})
    if request.method == 'POST':
        data = json.loads(request.body)
        if data['secret_code'] != os.getenv('FETCH_VOLUNTEER_NAMES_PWD'):
            return JsonResponse({'names': 'Oops! You seem to be an intruder!'})
        submitted = data['submitted']
        verified = data['verified']
        activity = data['activity']
        academic_year = data['academic_year']
        semester = data['semester']
        vols = Volunteer.objects.filter(submitted=submitted, verified=verified, activity=activity, registered_academic_year=academic_year, registered_semester=semester)
        names = []
        for v in vols:
            names.append(v.vname)
        return JsonResponse({'names':names})

def infoView(request):
    INFO_PWD2 = os.getenv('INFO_PWD2')
    if request.method == 'GET':
        return render (request, 'info_auth.html', {'INFO_PWD2' : INFO_PWD2})
    if request.method == 'POST':

        FETCH_VOLUNTEER_NAMES_PWD = os.getenv('FETCH_VOLUNTEER_NAMES_PWD')
        password = request.POST['password']
        academic_years_set = set()
        volunteers = Volunteer.objects.all()
        for v in volunteers:
            academic_years_set.add(str(v.registered_academic_year))

        total_not_yet_submitted = 0
        total_submitted_yet_to_be_verified = 0
        total_approved = 0
        total_rejected = 0
        total_failed_by_coords = 0
        total_total = 0
        total_failed_for_not_submitting_report = 0

        activities = Activity.objects.all()

        if password == os.getenv('INFO_PWD1'):
            curr = currentData.objects.get(index='Current')
            volunteerData = {}

            for act in activities:
                vols = Volunteer.objects.filter(activity = act.name, registered_semester = curr.Semester, registered_academic_year = curr.AcademicYear)
                not_yet_submitted = 0
                submitted_yet_to_be_verified = 0
                approved = 0
                rejected = 0
                failed_by_coords = 0
                total = 0
                failed_for_not_submitting_report = 0
                subData = {}
                for v in vols:
                    total += 1
                    total_total += 1
                    if v.submitted == 0 and v.verified == 0:
                        not_yet_submitted += 1
                        total_not_yet_submitted += 1
                    elif v.submitted == 1 and v.verified == 0:
                        submitted_yet_to_be_verified += 1
                        total_submitted_yet_to_be_verified += 1
                    elif v.submitted == 1 and v.verified == 1:
                        approved += 1
                        total_approved += 1
                    elif v.submitted == 0 and v.verified == 2:
                        rejected += 1
                        total_rejected += 1
                    elif v.submitted == 1 and v.verified == 3:
                        failed_by_coords += 1
                        total_failed_by_coords += 1
                    elif v.submitted == 0 and v.verified == 3:
                        failed_for_not_submitting_report += 1
                        total_failed_for_not_submitting_report += 1
                    subData = {
                        'Not Yet Submitted': not_yet_submitted,
                        'Submitted, yet to be verified' : submitted_yet_to_be_verified,
                        'Approved' : approved,
                        'Rejected, yet to fill again' : rejected,
                        'Failed for not meeting criteria' : failed_by_coords,
                        'Failed for not submitting report' : failed_for_not_submitting_report,
                        'Total Volunteers' : total
                            }
                volunteerData[act] = subData

            final_rep = {
                        "activities": activities,
                        "academic_years": list(academic_years_set),
                        'volunteerData':volunteerData,
                        "academic_year":curr.AcademicYear,
                        "sem":curr.Semester,
                        "default":True,
                        'total_not_yet_submitted' : total_not_yet_submitted,
                        'total_submitted_yet_to_be_verified' : total_submitted_yet_to_be_verified,
                        'total_approved' : total_approved,
                        'total_rejected' : total_rejected,
                        'total_failed_by_coords' : total_failed_by_coords,
                        'total_failed_for_not_submitting_report' : total_failed_for_not_submitting_report,
                        'total_total' : total_total,
                        'pwd' : FETCH_VOLUNTEER_NAMES_PWD,
                        'INFO_PWD2' : INFO_PWD2
                }
            return render (request, 'info_show.html', final_rep)


        elif password == os.getenv('INFO_PWD2'):
            type = request.POST['type']
            password = request.POST['password']
            chosenActivities = request.POST['activities'].split('.')
            chosenActivities.pop()
            academic_year = request.POST['academic_year']
            sem = request.POST['sem']


            if type == 'Volunteer':
                volunteerData = {}
                for act in chosenActivities:
                    vols = Volunteer.objects.filter(activity = act, registered_semester = sem, registered_academic_year = academic_year)

                    not_yet_submitted = 0
                    submitted_yet_to_be_verified = 0
                    approved = 0
                    rejected = 0
                    failed_by_coords = 0
                    total = 0
                    failed_for_not_submitting_report = 0
                    subData = {}
                    for v in vols:
                        total += 1
                        total_total += 1
                        if v.submitted == 0 and v.verified == 0:
                            not_yet_submitted += 1
                            total_not_yet_submitted += 1
                        elif v.submitted == 1 and v.verified == 0:
                            submitted_yet_to_be_verified += 1
                            total_submitted_yet_to_be_verified += 1
                        elif v.submitted == 1 and v.verified == 1:
                            approved += 1
                            total_approved += 1
                        elif v.submitted == 0 and v.verified == 2:
                            rejected += 1
                            total_rejected += 1
                        elif v.submitted == 1 and v.verified == 3:
                            failed_by_coords += 1
                            total_failed_by_coords += 1
                        elif v.submitted == 0 and v.verified == 3:
                            failed_for_not_submitting_report += 1
                            total_failed_for_not_submitting_report += 1
                        subData = {
                            'Not Yet Submitted' : not_yet_submitted,
                            'Submitted, yet to be verified' : submitted_yet_to_be_verified,
                            'Approved' : approved,
                            'Rejected, yet to fill again' : rejected,
                            'Failed for not meeting criteria' : failed_by_coords,
                            'Failed for not submitting report' : failed_for_not_submitting_report,
                            'Total Volunteers' : total
                                }
                    volunteerData[act] = subData

                final_rep = {
                    'activities' : activities,
                    'academic_years' : list(academic_years_set),
                    'volunteerData' : volunteerData,
                    'academic_year' : academic_year,
                    'sem' : sem,
                    'total_not_yet_submitted' : total_not_yet_submitted,
                    'total_submitted_yet_to_be_verified' : total_submitted_yet_to_be_verified,
                    'total_approved' : total_approved,
                    'total_rejected' : total_rejected,
                    'total_failed_by_coords' : total_failed_by_coords,
                    'total_failed_for_not_submitting_report' : total_failed_for_not_submitting_report,
                    'total_total' : total_total,
                    'pwd' : FETCH_VOLUNTEER_NAMES_PWD,
                    'INFO_PWD2' : INFO_PWD2
                    }
                return render (request, 'info_show.html', final_rep)



            if type == 'Coordinator':
                coordData = {}
                for act in chosenActivities:
                    vols = Volunteer.objects.filter(activity = act, registered_semester = sem, registered_academic_year = academic_year, submitted = 1, verified = 0)
                    c_yet_to_verify = {}
                    for v in vols:
                        if c_yet_to_verify.get(v.Cordinator) is not None:
                            c_yet_to_verify[v.Cordinator] += 1
                        else:
                            c_yet_to_verify[v.Cordinator] = 1
                    coordData[act] = dict(sorted(c_yet_to_verify.items()))
                return render (request, 'info_show.html', {"activities": activities, "academic_years": list(academic_years_set), 'coordData':coordData, "academic_year":academic_year, "sem":sem})
        else:
            messages.error(request, 'Incorrect Password')
            return redirect('info')



def sendEmailView(request):
    if request.method == 'GET':
        return render(request, 'send_email.html', {"form":FormWithCaptcha()})
    if request.method == 'POST':
        if not FormWithCaptcha(request.POST).is_valid():
            messages.error(request, 'Please verify that you are not a robot.')
            return redirect('send-email')
        if request.POST['secret_code'] != os.getenv('SEND_EMAIL_PWD'):
            messages.error(request, "Secret Code is wrong!")
            return redirect("send-email")
        emails = request.POST['emails'].strip().split(',')
        message = []
        str = ''
        for c in request.POST['messageArea']:
            if c == '\n':
                if str == '\n' or str == '' or str == '\r':
                    str = ''
                    continue
                message.append(str)
                str = ''
            else:
                str += c
        message.append(str)
        successMsg = ''
        for e in emails:
            successMsg += e + '\n'
            email_body = render_to_string("email_sending_template.html", {'messages':message})
            email = EmailMessage(request.POST['subject'], email_body, "noreply@semycolon.com", [e.strip()])
            email.content_subtype = "html"
            email.send(fail_silently=False)
        messages.success(request, 'Sent emails to \n\n' + successMsg)
        return redirect('send-email')

def Format_Name_Function(name):
    name = name.strip()
    if name == 'AnonymousUser' or len(name) < 3:
        return ''
    formatted_name = name[0].upper()
    i = 1
    while i < len(name):
        if name[i] == ' ':
            if name[i + 1] == ' ':
                i += 1
                continue
            formatted_name += ' ' + name[i + 1].upper()
            i += 1
        else:
            formatted_name += name[i].lower()
        i += 1
    return formatted_name

@csrf_exempt
def runFunction(request):
    if request.method == 'POST':
        if json.loads(request.body).get('secret_key') != os.getenv('RUN_FUNCTION_PWD'):
            return HttpResponse('Permission Denied !')

        Volunteer.objects.update(marked_IN_attendance=False)

        data = currentData.objects.get(index="Current")
        coordinators = Coordinator.objects.filter(registered_academic_year = data.AcademicYear, registered_semester = data.Semester)
        date = str(datetime.datetime.now().date().strftime("%d-%m-%Y"))

        stat = stats.objects.get(index=1)
        objs = Volunteer.objects.all()
        stat.vCount = len(objs)
        objs = Coordinator.objects.all()
        stat.cCount = len(objs)
        objs = Secretary.objects.all()
        stat.sCount = len(objs)
        stat.save()
        stat.uCount = stat.vCount + stat.cCount + stat.sCount
        stat.save()



        # Code to mail coords to verify reports
        vols = Volunteer.objects.filter(registered_academic_year = data.AcademicYear, registered_semester = data.Semester, submitted = 1, verified = 0)
        coords_to_mail = {}
        for v in vols:
            if v.Cordinator in coords_to_mail:
                coords_to_mail[v.Cordinator] += ', ' + v.vname
            else:
                coords_to_mail[v.Cordinator] = v.vname

        for coordName, volunteers in coords_to_mail.items():
            c = Coordinator.objects.get(cname=coordName)
            ss_report_verification = fe_report_verification = False

            if c.flagshipEvent != 'not_chosen' and c.flagshipEvent != '.' and c.flagshipEvent != '':
                FE_activityObj = Activity.objects.get(name=c.flagshipEvent)
                if FE_activityObj.report_verification:
                    fe_report_verification = True

            if c.activity != 'not_chosen' and c.activity != '.' and c.activity != '':
                SS_activityObj = Activity.objects.get(name=c.activity)
                if SS_activityObj.report_verification:
                    ss_report_verification = True

            if not ss_report_verification and not fe_report_verification:
                continue

            message = []
            message.append('Hello ' + c.cname + ',')
            message.append('Our software noticed that the below volunteer(s) have submitted their reports but you have not verified it yet. Please verify their reports (Approve/Reject/Fail it) as soon as possible!! The volunteer(s) are:')
            message.append(volunteers)
            message.append('Here\'s the login link for your convenience - https://swdc.pythonanywhere.com/a/login')
            message.append('Regards,')
            message.append('The Website Team')

            email_subject = 'URGENT: Please verify the Report(s) submitted to you!'
            email_body = render_to_string("email_sending_template.html", {'messages':message})
            email = EmailMessage(email_subject, email_body, "noreply@semycolon.com", [c.email])
            email.content_subtype = "html"
            email.send(fail_silently=False)





        try:
            with open('/home/swdc/SWDCWebsite/action/backup.txt', 'r+') as f:
                day = int(f.read())

                if day == 60:
                    file_path = '/home/swdc/SWDCWebsite/prod-database.db'
                    file_name = str(datetime.date.today()) + '-prod-database.db'
                    upload_to_google_drive(file_path, file_name, os.getenv('GDRIVE_UPLOAD_TOKEN'))
                    day = -1

                f.seek(0)
                f.write(str(day + 1))

        except Exception as e:
            print(e)






        vols = Volunteer.objects.filter(registered_academic_year = data.AcademicYear, registered_semester = data.Semester, submitted = 0, verified = 2)
        for v in vols:
            message = []
            message.append("Hi " + v.vname + ",")
            message.append("Your report was rejected by " + v.Cordinator)
            message.append("Reason: " + v.rejection_reason)
            message.append("Please edit and re-submit it immediately.")
            message.append('Here\'s the login link for your convenience - https://swdc.pythonanywhere.com/a/login')
            message.append("The Website Team")


            email_subject = 'URGENT: Your report was rejected'
            email_body = render_to_string("email_sending_template.html", {'messages':message})
            email = EmailMessage(email_subject, email_body, "noreply@semycolon.com", [v.email])
            email.content_subtype = "html"
            email.send(fail_silently=False)




        return HttpResponse('Done')
    else:
        return HttpResponse('Oops! You are a stranger here!')



def allotCoordinatorSequentialView(request):
    if request.method == "GET":
        activities = Activity.objects.all()
        return render(request, "allotment.html", {"activities": activities, "form":FormWithCaptcha()})
    else:
        if not FormWithCaptcha(request.POST).is_valid():
            messages.error(request, 'Please verify that you are not a robot.')
            return redirect('allot')
        entered_activity = request.POST["activity"]
        secret_code = request.POST["secret_code"]
        if secret_code != os.getenv('ALLOT_COORDS_SEQUENTIAL_PWD'):
            messages.error(request, "Incorrect secret code!")
            return redirect("allot")
        data = currentData.objects.get(index="Current")
        coordinators_doing_in_FE = Coordinator.objects.filter(
            flagshipEvent=entered_activity,
            registered_academic_year=data.AcademicYear,
            registered_semester=data.Semester,
        )
        coordinators_doing_in_SS = Coordinator.objects.filter(
            activity=entered_activity,
            registered_academic_year=data.AcademicYear,
            registered_semester=data.Semester,
        )
        coordinators = list(chain(coordinators_doing_in_FE, coordinators_doing_in_SS))
        volunteers = list(
            Volunteer.objects.filter(
                activity=entered_activity,
                registered_academic_year=data.AcademicYear,
                registered_semester=data.Semester,
            )
        )
        coordinator_count = len(coordinators)
        if coordinator_count == 0:
            messages.error(request, "There are no coordinators of " + entered_activity + " registered on website!")
            return redirect("allot")
        volunteer_count = len(volunteers)
        if volunteer_count == 0:
            messages.error(request, "There are no volunteers of " + entered_activity + " registered on website!")
            return redirect("allot")
        for i in range(0, volunteer_count):
            volunteers[i].Cordinator = coordinators[i % coordinator_count].cname
            volunteers[i].save()
        messages.success(request, "Coordinator allotment for " + entered_activity + "'s volunteers done successfully!")
        return redirect("allot")


def allotCoordsBySheet(request):
    if request.method == 'GET':
        return render(request, 'allot_by_sheet.html', {"form":FormWithCaptcha()})
    if request.method == 'POST':
        if not FormWithCaptcha(request.POST).is_valid():
            messages.error(request, 'Please verify that you are not a robot.')
            return redirect('allot-coords-by-sheet')
        if request.POST['secret_code'] != os.getenv('ALLOT_COORDS_SHEET_PWD'):
            messages.error(request, "Secret Code is wrong!")
            return redirect("allot-coords-by-sheet")

        uploaded_file = request.FILES['file']
        if 'xlsx' not in uploaded_file.name:
            messages.error(request, "Only files with a .xlsx extension are compatible.")
            return redirect("allot-coords-by-sheet")

        workbook = openpyxl.load_workbook(uploaded_file)
        sheet = workbook.active
        error = ''
        count = 0
        errCount = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            volunteer = Volunteer.objects.get(email = (row[3].strip()))
            try:
                coordName = row[5].strip()
                coord = Coordinator.objects.get(cname = coordName)
                volunteer.Cordinator = coord.cname
                count += 1
            except Exception as a:
                if row[5].strip() not in error:
                    error += '<br>' + row[5].strip()
                volunteer.Cordinator = 'Coord Not Found'
                errCount += 1
            volunteer.save()
        if error != '':
            msg = 'The following entries in the "Coordinator" column of the sheet could not be mapped with any coordinator currently registered on the website: <br>' + error
            messages.error(request, mark_safe(msg))
            messages.error(request,' Probably the coordinator(s) have registered with another name on the website and their name in the sheet does not match the name they have used to register on the website.')
            messages.error(request, 'The volunteers of the above coordinators have not been allotted a coordinator on the website. There are ' + str(errCount) + ' such volunteers.')
            messages.success(request, 'However, the remaining ' + str(count) + ' volunteers whose coordinator\'s names matched, were successfully allotted a coordiantor.')
            messages.error(request, 'Please enter the Coordinator\'s name in the sheet exactly as the name they have registered with and upload the sheet once again.')
        else:
            messages.success(request, 'Hurray! All the ' + str(count) + ' volunteers have been allotted a coordinator as per the sheet you uploaded !!')
        return redirect('allot-coords-by-sheet')


def downloadVolunteerReportView(request):
    academic_years_set = set()
    volunteers = Volunteer.objects.all()
    for v in volunteers:
        academic_years_set.add(str(v.registered_academic_year))
    if request.method == "GET":
        activities = Activity.objects.all()
        return render(request, "volunteer_data.html", {"activities": activities, "academic_years": list(academic_years_set), "form":FormWithCaptcha()})
    if request.method == "POST":
        if not FormWithCaptcha(request.POST).is_valid():
            messages.error(request, 'Please verify that you are not a robot.')
            return redirect('downloadVolunteerReportView')
        activityPOST = request.POST["activity"]
        ayPOST = request.POST["academic_year"]
        semPOST = request.POST["sem"]
        secret_code = request.POST["secret_code"]


        if secret_code != os.getenv('DOWNLOAD_VOL_DATA_PWD'):
            messages.error(request, "Secret Code is wrong!")
            return redirect("downloadVolunteerReportView")

        activity = []
        academicYear = []
        sem = []

        if activityPOST == 'ALL':
            acts = Activity.objects.all()
            for a in acts:
                activity.append(a.name)
        else:
            activity.append(activityPOST)

        if ayPOST == 'ALL':
            academicYear = list(academic_years_set)
        else:
            academicYear.append(ayPOST)

        if semPOST == 'ALL':
            sem.append(1)
            sem.append(2)
        else:
            sem.append(semPOST)

        rows = Volunteer.objects.filter(activity__in = activity, registered_academic_year__in = academicYear, registered_semester__in = sem)
        if len(rows) == 0:
            messages.error(request, "There are no volunteers of for the chosen fields.")
            return redirect("downloadVolunteerReportView")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = activityPOST
        ws.append([
            "Name",
            "Submitted",
            "Verified",
            "Email",
            "Gender",
            "Cordinator",
            "Guardian Faculty",
            "Activity",
            "Department",
            "Registered Academic Year",
            "Semester",
            "Academic Year",
            "Division",
            "Phone Number",
            "Parent's Phone Number",
            "Blood Group",
            "PRN",
            "Roll No.",
            "Current Address",
            "Objective Of The Activity",
            "Description Of The Activity",
            "Benefits To The Society",
            "Benefits To Self",
            "Learning Experiences & Challenges",
            "How Did It Shape Your Empathy",
            "URL",
            "Rejection Reason",
            "Attendance",
        ])

        for v in rows:
            ws.append([v.vname, v.submitted, v.verified, v.email, v.gender, v.Cordinator, v.guardian_faculty, v.activity, v.dept, v.registered_academic_year, v.registered_semester, v.academic_year, v.div, v.contact_num, v.parent_num, v.blood_group, v.prn, v.roll, v.current_add, v.ans1, v.ans2, v.ans3, v.ans4, v.ans5, v.ans6, v.url, v.rejection_reason, v.attendance])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        fname = 'Volunteers of Activity=' + str(activityPOST) + ' AY=' + str(ayPOST) + ' Sem=' + str(semPOST) + '.xlsx'
        response['Content-Disposition'] = f'attachment; filename={fname}'
        return response




def downloadCoordinatorReportView(request):
    if request.method == "GET":
        domains = Domain.objects.all()
        activities = Activity.objects.all()
        academic_years = set()
        coords = Coordinator.objects.all()
        for c in coords:
            academic_years.add(str(c.registered_academic_year))
        return render(
            request,
            "coord_data.html",
            {"domains": domains, "activities":activities, "academic_years": list(academic_years), "form":FormWithCaptcha()},
        )
    if request.method == "POST":
        if not FormWithCaptcha(request.POST).is_valid():
            messages.error(request, 'Please verify that you are not a robot.')
            return redirect('downloadCoordinatorReportView')
        domain = request.POST["domain"]
        activity = request.POST["activity"]
        filter_by = request.POST["filter_by"]
        academic_year = request.POST["academic_year"]
        sem = request.POST["sem"]
        secret_code = request.POST["secret_code"]

        if secret_code != os.getenv('DOWNLOAD_COORD_DATA_PWD'):
            messages.error(request, "Secret Code is wrong!")
            return redirect("downloadCoordinatorReportView")

        if filter_by == 'Activity':
            filename_secondPart = 'Activity - ' +activity
        else:
            filename_secondPart = 'Domain - ' + domain



        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = filename_secondPart
        ws.append([
            "Name",
            "Submitted",
            "Verified",
            "Email",
            "Gender",
            "Secretary",
            "Social Actvity",
            "Flagship Event",
            "Domain",
            "Department",
            "Registered Academic Year",
            "Semester",
            "Academic Year",
            "Division",
            "Phone Number",
            "Parent's Phone Number",
            "Blood Group",
            "PRN",
            "Roll No.",
            "Current Address",
            "Objective Of The Activity",
            "Description Of The Activity",
            "Benefits To The Society",
            "Benefits To Self",
            "Learning Experiences & Challenges",
            "How Did It Shape Your Empathy",
            "URL",
        ])


        if filter_by == 'Domain':
            rows = Coordinator.objects.filter(domain=domain, registered_academic_year=academic_year, registered_semester=sem)
        else:
            if(Activity.objects.get(name = activity).flagship_event):
                rows = Coordinator.objects.filter(flagshipEvent = activity, registered_academic_year = academic_year, registered_semester = sem)
            else:
                rows = Coordinator.objects.filter(activity = activity, registered_academic_year=academic_year, registered_semester=sem)


        if not rows.exists():
            messages.error(request, "We found no records for the choosen fields")
            return redirect("downloadCoordinatorReportView")

        for c in rows:
            ws.append([
                c.cname,
                c.submitted,
                c.verified,
                c.email,
                c.gender,
                c.Secretary,
                c.activity,
                c.flagshipEvent,
                c.domain,
                c.dept,
                c.registered_academic_year,
                c.registered_semester,
                c.academic_year,
                c.div,
                c.contact_num,
                c.parent_num,
                c.blood_group,
                c.prn,
                c.roll,
                c.current_add,
                c.ans1,
                c.ans2,
                c.ans3,
                c.ans4,
                c.ans5,
                c.ans6,
                c.url
                ])


        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        fname = 'Coordinator - ' + filename_secondPart + '.xlsx'
        response['Content-Disposition'] = f'attachment; filename={fname}'

        return response



def testCertificateView(request):
    if request.method == "GET":
        return render (request, 'test_certificate.html', {"form":FormWithCaptcha()})
    if request.method == "POST":
        if not FormWithCaptcha(request.POST).is_valid():
            messages.error(request, 'Please verify that you are not a robot.')
            return redirect('test-certificate')
        if request.POST['secret_code'] != os.getenv('TEST_CERTIFICATE_PWD'):
            messages.error(request, "Incorrect secret code!")
            return redirect("test-certificate")
        certificateFile = request.FILES['uploaded_name']
        packet = io.BytesIO()
        output_buffer = io.BytesIO()
        canvasObj = canvas.Canvas(packet, pagesize=landscape(A3))
        canvasObj.setFont("Times-Bold", 27)
        width, height = A3
        text_width = canvasObj.stringWidth('Test Name', "Times-Roman", 27)
        x_center = (width - text_width) / 2
        canvasObj.drawString(x_center, 330, 'Test Name')
        canvasObj.save()
        packet.seek(0)
        new_pdf = PdfReader(packet)
        certificate_template = PdfReader(certificateFile)
        output = PdfWriter()
        page = certificate_template.pages[0]
        page.merge_page(new_pdf.pages[0])
        output.add_page(page)
        output.write(output_buffer)
        output_buffer.seek(0)
        response = HttpResponse(output_buffer, content_type='application/pdf')
        return response


def downloadVolunteerCertificatesZip(request):
    activities = Activity.objects.all()
    academic_years = set()
    volunteers = Volunteer.objects.all()
    for v in volunteers:
        academic_years.add(str(v.registered_academic_year))
    if request.method == "GET":
        return render(request, "download-volunteer-certificates-zip.html", {"activities": activities, 'academic_years':academic_years, "form":FormWithCaptcha()})
    if request.method == 'POST':
        if not FormWithCaptcha(request.POST).is_valid():
            messages.error(request, 'Please verify that you are not a robot.')
            return redirect('download-volunteer-certificates-zip')

        if request.POST["secret_code"] != os.getenv('DOWNLOAD_VOL_CERTIFICATES_ZIP_PWD'):
            messages.error(request, "Incorrect secret code!")
            return redirect("download-volunteer-certificates-zip")

        act = request.POST["act"].replace(' ', '_')
        year = request.POST["year"]
        semester = request.POST["semester"]

        volunteers = Volunteer.objects.filter(activity = request.POST["act"], registered_academic_year = year, registered_semester = semester, verified = 1)
        if len(volunteers) == 0:
            messages.error(request, 'No volunteers have cleared ' + act + ' in AY ' + year + ' and semester ' + semester)
            return redirect("download-volunteer-certificates-zip")
        template_path = os.path.join(settings.BASE_DIR, 'certificateTemplates') + act + '/' + act + '_' + year + '_' + semester + ".pdf"
        if not os.path.exists(template_path):
            messages.error(request, 'The certificate template for ' + act + ' for AY ' + year + ' and sem ' + semester + ' is not available.')
            return redirect("download-volunteer-certificates-zip")
        os.makedirs('temp')
        temp_path = 'temp/'
        for volunteer in volunteers:
            packet = io.BytesIO()
            canvasObj = canvas.Canvas(packet, pagesize=landscape(A3))
            canvasObj.setFont("Times-Bold", 27)
            width, height = A3
            text_width = canvasObj.stringWidth(volunteer.vname, "Times-Bold", 27)
            x_center = (width - text_width) / 2
            canvasObj.drawString(x_center, 330, volunteer.vname)
            canvasObj.save()
            packet.seek(0)
            new_pdf = PdfReader(packet)
            certificate_template = PdfReader(open(template_path, "rb"))
            output = PdfWriter()
            page = certificate_template.pages[0]
            page.merge_page(new_pdf.pages[0])
            output.add_page(page)
            pdf_file_path = temp_path + volunteer.vname + ".pdf"
            output_stream = open(pdf_file_path, "wb")
            output.write(output_stream)
            output_stream.close()
        pdf_files = []
        for filename in os.listdir(temp_path):
            full_path = os.path.join(temp_path, filename)
            if filename.endswith(".pdf"):
                pdf_files.append(full_path)
        output_zip = (act + "_" + year + "_Sem_" + semester + ".zip")
        zipf = zipfile.ZipFile(output_zip, "w", zipfile.ZIP_DEFLATED)
        for pdf_file in pdf_files:
            try:
                zipf.write(pdf_file, arcname=os.path.basename(pdf_file))
                os.remove(pdf_file)
            except Exception as e:
                print(f"Error adding {pdf_file} to ZIP: {str(e)}")
        zipf.close()
        with open(output_zip, "rb") as zip_file:
            response = HttpResponse(zip_file.read(), content_type="application/zip")
            response["Content-Disposition"] = f"attachment; filename={output_zip}"
        os.remove(output_zip)
        os.rmdir(temp_path)
        return response


def generateIndividualCertificate(request):
    if request.method == 'GET':
        academic_years = set()
        volunteers = Volunteer.objects.all()
        for v in volunteers:
            academic_years.add(str(v.registered_academic_year))
        activities = Activity.objects.all()
        return render (request, 'generate-individual-certificate.html', {"activities": activities, "form":FormWithCaptcha(), "academic_years" : academic_years})
    if request.method == 'POST':
        if not FormWithCaptcha(request.POST).is_valid():
            messages.error(request, 'Please verify that you are not a robot.')
            return redirect('generate-individual-certificate')
        if request.POST["secret_code"] != os.getenv('GENERATE_INDV_CERTIFICATE_PWD'):
            messages.error(request, "Incorrect secret code!")
            return redirect("generate-individual-certificate")
        name = request.POST["name"]
        act = request.POST["act"].replace(" ", "_")
        year = request.POST['year']
        sem = request.POST['sem']

        template_path = os.path.join(settings.BASE_DIR, 'certificateTemplates') + act + '/' + act + '_' + year + '_' + sem + '.pdf'
        if not os.path.exists(template_path):
            messages.error(request, 'The certificate template for ' + act + ' for AY ' + year + ' and sem ' + sem + ' is not available.')
            return redirect("generate-individual-certificate")
        certificate_template = PdfReader(open(template_path, "rb"))
        packet = io.BytesIO()
        output_buffer = io.BytesIO()
        canvasObj = canvas.Canvas(packet, pagesize=landscape(A3))
        canvasObj.setFont("Times-Bold", 27)
        width, height = A3
        text_width = canvasObj.stringWidth(name, "Times-Roman", 27)
        x_center = (width - text_width) / 2
        canvasObj.drawString(x_center, 330, name)
        canvasObj.save()
        packet.seek(0)
        new_pdf = PdfReader(packet)
        output = PdfWriter()
        page = certificate_template.pages[0]
        page.merge_page(new_pdf.pages[0])
        output.add_page(page)
        output.write(output_buffer)
        output_buffer.seek(0)
        response = HttpResponse(output_buffer, content_type='application/pdf')
        return response


def failVolunteerView(request):
    if request.method == "GET":
        activities = Activity.objects.all()
        academic_years = set()
        volunteers = Volunteer.objects.all()
        for v in volunteers:
            academic_years.add(str(v.registered_academic_year))
        return render(request, 'fail-volunteers.html', {'activities': activities, 'academic_years':list(academic_years), "form":FormWithCaptcha()})

    if request.method == 'POST':
        if not FormWithCaptcha(request.POST).is_valid():
            messages.error(request, 'Please verify that you are not a robot.')
            return redirect('fail')
        secret_code = request.POST['secret_code']
        if secret_code != os.getenv('FAIL_VOL_PWD'):
            messages.error(request, "Incorrect secret code!")
            return redirect("fail")
        activity = request.POST['activity']
        academic_year = request.POST['academic_year']
        semester = request.POST['semester']
        failed_volunteers = Volunteer.objects.filter(activity = activity, registered_academic_year = academic_year,registered_semester = semester, submitted = 0, verified = 2)
        if len(failed_volunteers) == 0:
            messages.error(request, 'There are no volunteers to fail for the selected fields.')
            return redirect('fail')

        email_subject = "Social Services Course: Update"
        formatedMsg = []
        formatedMsg.append("The 'Social Services' course required your active participation and successful completion of " + activity + " being the primary criterion for passing.")
        formatedMsg.append("Timely submission of the report was mandatory as per assessment guidelines, ensuring successful progression in the activity.")
        formatedMsg.append("Unfortunately, the deadline for the report filling process, which was set for 21 November at 6 pm, was not met in your case.")
        formatedMsg.append("Regrettably, due to the missed deadline, we must inform you that you've failed in the 'Social Services' course.")
        # formatedMsg.append("The 'Social Service' course required your active participation in Utkarsh, a main criterion for successful completion. Regrettably, our prior email notification regarding the course completion was sent to you due to a technical glitch.")
        # formatedMsg.append("It has come to our attention that you haven't engaged in any sessions among the many allotted ones. Furthermore, there has been no communication initiated with the activity head or coordinator to explicate the reasons for your non-participation.")
        # formatedMsg.append("Consequently, we must convey that, owing to these circumstances, you have been marked as Failed in the 'Social Service' course.")

        for v in failed_volunteers:
            v.submitted = 1
            v.verified = 3
            v.rejection_reason = 'Not submitted report before 6pm, 21st November, 2023.'
            v.save()
            context = {"name": v.vname , "messages": formatedMsg}
            email_body = render_to_string("email_template.html", context)
            email = EmailMessage(
                email_subject,
                email_body,
                "noreply@semycolon.com",
                [v.email],
            )
            email.content_subtype = "html"
            email.send(fail_silently=False)
        messages.success(request, "The selected volunteers have been failed and mailed about it.")
        return redirect("fail")

def report_data(request):
    activities = Activity.objects.all()
    current = currentData.objects.get(index='Current')
    activity_data = []

    for activity in activities:
        volunteers = Volunteer.objects.filter(
            activity=activity,
            registered_academic_year=current.AcademicYear,
            registered_semester=current.Semester
        )

        not_yet_submitted = sum(1 for v in volunteers if v.submitted == 0 and v.verified == 0)
        submitted_yet_to_be_verified = sum(1 for v in volunteers if v.submitted == 1 and v.verified == 0)
        verified = sum(1 for v in volunteers if v.submitted == 1 and v.verified == 1)
        rejected = sum(1 for v in volunteers if v.submitted == 0 and v.verified == 2)
        failed_by_coords = sum(1 for v in volunteers if v.submitted == 1 and v.verified == 3)
        failed_for_not_submitting_report = sum(1 for v in volunteers if v.submitted == 0 and v.verified == 3)
        total = volunteers.count()

        activity_data.append({
            'name': activity.name,
            'not_yet_submitted': not_yet_submitted,
            'submitted_yet_to_be_verified': submitted_yet_to_be_verified,
            'verified': verified,
            'rejected': rejected,
            'failed_by_coords': failed_by_coords,
            'failed_for_not_submitting_report': failed_for_not_submitting_report,
            'total': total
        })

    return render(request, "reportSubmission.html", {'activities': activity_data})
